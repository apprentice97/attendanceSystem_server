import json
import os
import time
import numpy as np
import openpyxl
from django.conf import settings
from django.contrib.sites import requests
from django.db import transaction
from django.http import JsonResponse, HttpResponse

from common.models import Teacher, Course, TeacherCourse, Attendance, StudentCourse, Student
from myMethod import send_mail

# CRUD（create, read, update, delete）


def dispatcher(request):
    if request.method == 'GET':
        request.params = request.GET
    elif request.method == 'POST':
        request.params = request.POST
    action = request.params['action']
    print(request.params)
    print(request.body)
    # elif request.method in ['POST', 'PUT', 'DELETE']:
    # # 根据接口，POST/PUT/DELETE 请求的消息体都是 json格式
    # request.params = json.loads(request.body)
    if action == 'teacher_login':
        return teacher_login(request)
    elif action == 'list_course':
        return list_course(request)
    elif action == 'course_information':
        return course_information(request)
    elif action == 'course_attendance':
        return course_attendance(request)
    elif action == 'course_attendance_information':
        return course_attendance_information(request)
    elif action == 'call_name':
        return call_name(request)
    elif action == "teacher_get_self_info":
        return teacher_get_self_info(request)
    elif action == "list_application_for_leave":
        return list_application_for_leave(request)
    elif action == 'reject_application':
        return reject_application(request)
    elif action == 'approve_application':
        return approve_application(request)
    elif action == 'teacher_modify_attendance':
        return teacher_modify_attendance(request)
    elif action == 'teacher_request_statistics':
        return teacher_request_statistics(request)
    elif action == 'request_image':
        return request_image(request)
    elif action == 'teacher_reset_password':
        return teacher_reset_password(request)
    else:
        return JsonResponse({'ret': 400, 'msg': '不支持该类型http请求！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def teacher_login(request):
    try:
        student_id = request.params['account']
        password = request.params['password']
        try:
            student = Teacher.objects.get(teacher_id=student_id)
        except Teacher.DoesNotExist:
            return JsonResponse({'ret': 2}, json_dumps_params={'ensure_ascii': False})
        if password == student.password:
            return JsonResponse({'ret': 0}, json_dumps_params={'ensure_ascii': False})
        else:
            return JsonResponse({'ret': 1}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def list_course(request):
    try:
        teacher_id = request.params['account']
        teacher_courses = TeacherCourse.objects.filter(teacher_id=teacher_id).values()
        ret = []
        for teacher_course in teacher_courses:
            for key, value in teacher_course.items():
                if "course_id" == key:
                    courses = Course.objects.filter(course_id=value)
                    for course in courses:
                        ret.append(course.json())
        return JsonResponse({'ret': 0, 'data': ret}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def course_information(request):
    try:
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        attendances = Attendance.objects.filter(teacher_id=teacher_id, course_id=course_id)
        ret = []
        for attendance in attendances:
            ret.append(attendance.json())
        return JsonResponse({'ret': 0, 'data': ret}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})



'''
    type == 0 出勤
    type == 1 病假
    type == 2 事假
    type == 3 缺勤
    modify == 0 或者 modify == 2 默认为缺勤
'''


@transaction.atomic
def course_attendance(request):
    try:
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        attendances = Attendance.objects.filter(teacher_id=teacher_id, course_id=course_id)
        arr = np.zeros((1000, 5))
        ret = []
        max_sn = 0
        my_dict = dict()
        for attendance in attendances:
            serial_number = int(attendance.serial_number)
            number = int(attendance.type)
            time_str = attendance.time
            modify = int(attendance.modify)
            if 0 == modify:
                arr[serial_number][3] = arr[serial_number][3] + 1
            elif 2 == modify:
                arr[serial_number][3] = arr[serial_number][3] + 1
            elif 0 == number:
                arr[serial_number][0] = arr[serial_number][0] + 1
            elif 1 == number:
                arr[serial_number][1] = arr[serial_number][1] + 1
            elif 2 == number:
                arr[serial_number][2] = arr[serial_number][2] + 1
            elif 3 == number:
                arr[serial_number][3] = arr[serial_number][3] + 1
            arr[serial_number][4] = arr[serial_number][4] + 1
            my_dict[serial_number] = time_str
            max_sn = max(max_sn, serial_number)
        for i in range(1, max_sn + 1):
            ret.append({'序号': i, '时间': my_dict[i], '出勤': str(int(arr[i][0])), '病假': str(int(arr[i][1])),
                        '事假': str(int(arr[i][2])), '缺勤': str(int(arr[i][3])), '应到': str(int(arr[i][4]))})
        return JsonResponse({'ret': 0, 'data': ret},
                            json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def course_attendance_information(request):
    try:
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        serial_number = request.params['serial_number']
        attendances = Attendance.objects.filter(teacher_id=teacher_id, course_id=course_id, serial_number=serial_number).order_by("student_id")
        ret = []
        for attendance in attendances:
            ret.append(attendance.json())
        return JsonResponse({'ret': 0, 'data': ret}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def call_name(request):
    try:
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        student_ids = []
        serial_number = 0

        student_courses = StudentCourse.objects.filter(course_id=course_id)
        for student_course in student_courses:
            student_ids.append(student_course.student.student_id)
        attendances = Attendance.objects.filter(course_id=course_id)
        for attendance in attendances:
            serial_number = max(attendance.serial_number, serial_number)
        serial_number = serial_number + 1
        time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        for student_id in student_ids:
            Attendance.objects.create(serial_number=serial_number,
                                      student_id=student_id,
                                      course_id=course_id,
                                      teacher_id=teacher_id,
                                      modify=0,
                                      type=0,
                                      time=time_str)
        return JsonResponse({'ret': 0, 'data': student_ids}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def teacher_get_self_info(request):
    teacher_id = request.params['teacher_id']
    try:
        teacher = Teacher.objects.get(teacher_id=teacher_id)
    except Teacher.DoesNotExist:
        return JsonResponse({'ret': -1, "msg": "该学生不存在"}, json_dumps_params={'ensure_ascii': False})
    return JsonResponse({'ret': 0, 'teacher_name': teacher.name, 'teacher_email': teacher.email},
                        json_dumps_params={'ensure_ascii': False})


@transaction.atomic
def list_application_for_leave(request):
    print("list_application_for_leave")
    print(request.params)
    try:
        teacher_id = request.params['teacher_id']
        attendances = Attendance.objects.filter(teacher_id=teacher_id, modify=2).order_by("time", "student_id")
        ret = []
        for attendance in attendances:
            ret.append(attendance.json())
        return JsonResponse({'ret': 0, 'data': ret}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic()
def reject_application(request):
    print("reject_application")
    print(request.params)
    try:
        student_id = request.params['student_id']
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        serial_number = request.params['serial_number']
        try:
            attendance = Attendance.objects.get(student_id=student_id, course_id=course_id, teacher_id=teacher_id,
                                                serial_number=serial_number)
        except Attendance.DoesNotExist:
            return JsonResponse({'ret': 1, 'data': '该记录不存在!'}, json_dumps_params={'ensure_ascii': False})
        attendance.modify = 1
        attendance.type = 3
        attendance.save()
        return JsonResponse({'ret': 0, 'data': '修改成功! '}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic()
def approve_application(request):
    try:
        student_id = request.params['student_id']
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        serial_number = request.params['serial_number']
        try:
            attendance = Attendance.objects.get(student_id=student_id, course_id=course_id, teacher_id=teacher_id,
                                                serial_number=serial_number)
        except Attendance.DoesNotExist:
            return JsonResponse({'ret': 1, 'data': '该记录不存在!'}, json_dumps_params={'ensure_ascii': False})
        attendance.modify = 1
        attendance.save()
        return JsonResponse({'ret': 0, 'data': '修改成功! '}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


@transaction.atomic()
def teacher_modify_attendance(request):
    try:
        student_id = request.params['student_id']
        course_id = request.params['course_id']
        teacher_id = request.params['teacher_id']
        serial_number = request.params['serial_number']
        my_type = request.params['type']
        try:
            attendance = Attendance.objects.get(student_id=student_id, course_id=course_id, teacher_id=teacher_id,
                                                serial_number=serial_number)
        except Attendance.DoesNotExist:
            return JsonResponse({'ret': 1, 'data': '该记录不存在!'}, json_dumps_params={'ensure_ascii': False})
        attendance.modify = 1
        attendance.type = int(my_type)
        attendance.save()
        return JsonResponse({'ret': 0, 'data': '修改成功! '}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


def teacher_request_statistics(request):
    try:
        teacher_id = request.params['teacher_id']
        teacher_email =request.params['teacher_email']
        attendances = Attendance.objects.filter(teacher_id=teacher_id).\
            order_by("course_id", "serial_number", "student_id")

        student_dict = {'student_id': 'student_name'}
        students = Student.objects.filter()
        for student in students:
            student_dict[student.student_id] = student.name
        book = openpyxl.Workbook()
        last_serial_number = 0
        last_course_id = "0"
        row = None
        sh = None
        type_dict = {0: "出勤", 1: "病假", 2: "事假", 3: "缺勤"}
        for attendance in attendances:
            if last_course_id != attendance.course_id:
                last_course_id = attendance.course_id
                sh = book.create_sheet(attendance.course_id)
                sh['A1'] = '学号'
                sh['B1'] = '姓名'
            if last_serial_number != attendance.serial_number:
                last_serial_number = attendance.serial_number
                sh.cell(1, last_serial_number + 2).value = "第" + str(last_serial_number) + "次"
                row = 2
            sh.cell(row, 1).value = attendance.student_id
            sh.cell(row, 2).value = student_dict[attendance.student_id]
            if attendance.modify == 0 or attendance.modify == 2:
                sh.cell(row, last_serial_number + 2).value = type_dict[3]
            else:
                sh.cell(row, last_serial_number + 2).value = type_dict[attendance.type]
            row = row + 1
        book.remove(book['Sheet'])
        book.save(teacher_id + '.xlsx')
        send_mail(teacher_email, teacher_id + '.xlsx')
        return JsonResponse({'ret': 0, 'data': '请求成功! '}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        print(e)
        return JsonResponse({'ret': 500, 'msg': '服务端发生异常！'}, json_dumps_params={'ensure_ascii': False})


def request_image(request):
    try:
        image_path = os.path.join(settings.BASE_DIR,
                                 "resources\\picture\\attendance\\" + request.params['picture_name'] + ".png")
        with open(image_path, 'rb') as f:
            image_data = f.read()
        return HttpResponse(image_data, content_type="image/png")
    except Exception as e:
        print(e)
        return HttpResponse(str(e))


def teacher_reset_password(request):
    teacher_id = request.params['teacher_id']
    password = request.params['password']
    try:
        teacher = Teacher.objects.get(teacher_id=teacher_id)
        teacher.password = password
        teacher.save()
    except Student.DoesNotExist:
        return JsonResponse({'ret': -1, "msg": "该学生不存在"}, json_dumps_params={'ensure_ascii': False})
    return JsonResponse({'ret': 0}, json_dumps_params={'ensure_ascii': False})